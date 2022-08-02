
##UPDATES - 2022 May 02 17:08
##- super fast! No unecessary print statements
##- Investors automatically allocated with additional note for W CHEN
##- (investors.txt)
##- fixed infini loop for note added
##- changed Gale Close colour
##- addNoteQ function added
##- (4.0.1) can choose to add note to autonotes



import datetime
from enum import auto
from tkinter import *
#from tkinter import filedialog
import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl import load_workbook
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
import pandas as pd
import os

import files


columns = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J"]

global accountSaveFiles, roomSaveFiles, accountSaveStrings, accountSaveStringsLong, accountNamesString, accountNames, accountColoursFill, Memo, Note, none, NotesNote, NotesMemo
accountSaveFiles, roomSaveFiles, accountSaveStrings, accountSaveStringsLong, accountNamesString, accountNames, accountColoursFill, Memo, Note, none, NotesNote, NotesMemo, investors = files.getAll("Saves", "Rooms")

# --------------------------------------------------=======[[[[[[end of dictionary for account names/numbers]]]]]]=======---------------------------------------------------

def update():
  accountSaveFiles, roomSaveFiles, accountSaveStrings, accountSaveStringsLong, accountNamesString, accountNames, accountColoursFill, Memo, Note, none, NotesNote, NotesMemo, investors = files.getAll("Saves", "Rooms")

def conversion(n, x):
  for i in range(0, n):
    # takes (x[i]) out of list
    hi = x[i]
##    print(hi)

    # splits it into another list
    y = hi.split(",") # creates list
##    print(y)

    #replaces non-list (hi) with (hi) as a list
    x.pop(i)
    x.insert(i, y)
##  print(x)
##  print("\n\n\n\n")
  return x

def itr():
  global TyrellR, IllingworthR, Gale_Close39R, SilburyR, Flanders54R, Mount27R, SheldrickR, North_PlaceR, GarendonR, roomSaveFiles
  for i in range(0, len(roomSaveFiles)):
    roomSaveFiles[i] = conversion(len(roomSaveFiles[i]), roomSaveFiles[i])
##  print(roomSaveFiles)


def addNoteQ(lineArr, lineCount):
  notesCheck = False
  while notesCheck == False:
    noteAdd = input("Do you want to add a note?\nYes/y or No/n: ")
    if noteAdd == "Yes" or noteAdd == "yes" or noteAdd == "y":
      note = input("What note do you want to add?\nNote: ")
      cellAdd = "J" + str(lineCount)
      ws[cellAdd] = note
      ws["J" + str(lineCount)].fill = none
      print("Added")
      notesCheck = True
      
      autoNotesCheck = True
      while autoNotesCheck:
        autoNoteAdd = input("Do you want to add this as an auto note?\nYes/y or No/n: ").lower()
        if autoNoteAdd == "yes" or autoNoteAdd == "y":
          Memo.append(lineArr[5])
          file = open(os.path.join("Notes", "NotesMemo.txt"), "w")
          for i in range(0, len(Memo)):
            if i == len(Memo) - 1:
              file.write(Memo[i])
            else:
              file.write(Memo[i] + "\n")
          file.close()
          Note.append(note)
          file = open(os.path.join("Notes", "NotesNote.txt"), "w")
          for i in range(0, len(Note)):
            if i == len(Note) - 1:
              file.write(Note[i])
            else:
              file.write(Note[i] + "\n")
          file.close()
          autoNotesCheck = False
        elif autoNoteAdd == "no" or autoNoteAdd == "n":
          print("OK")
          autoNotesCheck = False
        else:
          print("Sorry. I didn't understand that")
      break
    elif noteAdd == "No" or noteAdd == "no" or noteAdd == "n":
      print("OK")
      notesCheck = True
    else:
      print("Sorry. I didn't understand that")

def display_lists():
  print("\nTyrell: " + str(Tyrell))
  print("\n\nIllingworth: " + str(Illingworth))
  print("\n\nGale_Close39: " + str(Gale_Close39))
  print("\n\nSilbury: " + str(Silbury))
  print("\n\nFlanders54: " + str(Flanders54))
  print("\n\nMount27: " + str(Mount27))
  print("\n\nSheldrick: " + str(Sheldrick))
  print("\n\nNorth_Place: " + str(North_Place))
  print("\n\nGarendon: " + str(Garendon))



def readColumnF(ws):
  lineArr = []
  lineCount = 0
  foundVal = False
  ws["G1"] = "Property"
  ws["H1"] = "Room Number"
  ws["I1"] = "Auto Notes"
  ws["J1"] = "Notes"

  for row in ws.iter_rows(min_row = 1, max_row = linesInSheet, max_col = 6, values_only = True):
    foundVal = False
    lineCount += 1
    global Tyrell, Illingworth, Gale_Close39, Silbury, Flanders54, Mount27, Sheldrick, North_Place, Garendon

    for investor in investors:
      if investor in row[5]:
        for Column in range(1, 8):
          rangeColour = columns[Column] + str(lineCount)
          ws[rangeColour].fill = none
        print("Investor spotted: ", row[5])
        if investor == "W CHEN" and "PROFIT" not in row[5]:
          addNoteQ(row, lineCount)
        foundVal = True
        break
      
    for propName in range(0, 9):
      propSearchTerm = len(accountNames[propName])
      for eachPropSearchTerm in range(0, propSearchTerm):
        searchVal = accountNames[propName][eachPropSearchTerm]
        if searchVal == '':
          print("")
        else:
          lineArr = row
          if lineArr[5] is None:
            print("Error reading line F", str(lineCount) + ".", "Please fix the problem and save the excel file. Then restart the program.", "")
          elif (searchVal in lineArr[5]) and (foundVal == False):
            checkMemo(lineArr[5], lineCount)
            checkRoom(lineArr[5], lineCount, propName)
            cell = "G" + str(lineCount)
            ws[cell] = accountNamesString[propName]
            colourIn(lineCount, propName)
            foundVal = True
              

    if foundVal == False:
      if lineCount != 1:
        print("\n\n\nNo matching values found for line", lineCount, "with date", str(lineArr[1]).split("00:")[0])
        cell = "F" + str(lineCount)
        print("|" + ws[cell].value + "| is what it says")
        foundVal = False
        while foundVal == False:
          print("Please manually input the property(or type |none| if there is no property)\n" + str(accountNamesString))
          manualInput = input("User: ")
          foundPropName = False
          for propName in range(0, 9):
            if manualInput == accountNamesString[propName]:
              foundPropName = True
              foundVal = True
              cell = "G" + str(lineCount)
              ws[cell] = accountNamesString[propName]
              colourIn(lineCount, propName)
              print("\nFound match for line", lineCount, "")
              print("\n\nWould you like to add that Tenant Name/Company ID?")
              didAns = False
              while didAns == False:
                Addition = input("Yes/y or No/n: ")
                if Addition == "No" or Addition == "no" or Addition == "n":
                  print("\nGot it\n")
                  didAns = True
                elif Addition == "Yes" or Addition == "yes" or Addition == "y":
                  didAns = True
                  update()
                  check = False
                  while check == False:
                    propChange = manualInput + "F"
                    if propChange in accountSaveStringsLong:
                      print("Found")
                      numberOfSaves = len(accountSaveFiles)
                      for i in range(0, numberOfSaves):
                        if propChange == accountSaveStrings[i]:
                          addSearchTerm(accountSaveFiles[i], i)
                          print("\nDone\n\n")
                          check = True
              addNoteQ(lineArr, lineCount)
              update()

            if manualInput == "none" and foundPropName == False:
              for Column in range(1, 8):
                rangeColour = columns[Column] + str(lineCount)
                ws[rangeColour].fill = none

              addNoteQ(lineArr, lineCount)
              update()

              foundPropName = True
              foundVal = True
          if foundPropName == False:
            print("Sorry, I am unable to find", manualInput + ".", "Please type the property exactly like these:\n", accountNamesString, "or 'none'", "")

def checkMemo(colF, lineCount):
  cell = "I" + str(lineCount)
  for objct in range(0, len(Memo)):
    if Memo[objct] in colF:
      ws[cell] = Note[objct]
      
def checkRoom(colF, lineCount, propName):
  cell = "H" + str(lineCount)
  if ("Payprop" not in colF) or ("PayProp" not in colF) or ("PAYPROP" not in colF):
    for room in range(0, len(roomSaveFiles[propName])):
      if (roomSaveFiles[propName][room][0] in colF) and (roomSaveFiles[propName][room][0] != none):
        ws[cell] = roomSaveFiles[propName][room][1]
  elif ("Payprop" in colF) or ("PayProp" in colF) or ("PAYPROP" in colF):
    print("Payprop property. Which room is |" + str(colF) + "| in? (just number, e.g. 2)")
    roomNum = input("Room number: ") - 1
    isINT = isinstance(roomNum, int)
    ws[cell] = roomSaveFiles[propName][roomNum][1]
  else:
    print("No room for |" + str(colF) + "|")
    


def findPositive(lines):
  whereLook = "D" + str(lines)
  cellVal = ws.cell(row = lines, column = 4)
  try:
    int(cellVal.value)
  except:
    verify = True
    while verify == True:
      print("Box " + whereLook + " has value |" + str(cellVal.value) + "| Should the whole line be coloured in?")
      inputi = input("Yes/y or No/n: ")
      verify = False
      if inputi == "Yes" or inputi == "y":
        verify = False
        return True
      elif inputi == "No" or inputi == "n":
        verify = False
        return False
      else:
        verify = True
        print("Sorry. I do not understand. Please try again.")
        
  if int(cellVal.value) > 0:
    return True
  elif int(cellVal.value) < 0:
    return False

        
def colourIn(lineCount, propName):
  posOrNot = findPositive(lineCount)
  if posOrNot == True:
    for Column in range(1, 9):
      rangeColour = columns[Column] + str(lineCount)
      ws[rangeColour].fill = accountColoursFill[propName]
  elif posOrNot == False:
    rangeColour = "G" + str(lineCount)
    ws[rangeColour].fill = accountColoursFill[propName]
  
  
def replaceSearchTerm(fileRaw, i):
  file = open(fileRaw, "r")
  
  reading = file.read()
  lists = reading.split("\n")
  file.close()
  print(lists)
  numberInLists = len(lists)
  global Tyrell, Illingworth, Gale_Close39, Silbury, Flanders54, Mount27, Sheldrick, North_Place, Garendon
  found = False
  while found == False:
    print("\n\n\nWhat would you like to replace? (look in the Save file to find which one)")
    searchReplace = input("Tenant Name/Company ID: ")
    whatReplace = input("\nWhat would you like to replace it with?\nTenant Name/Company ID: ")
    if searchReplace in lists:
      for i in range(0, int(numberInLists)):
        if lists[i] == searchReplace:
          lists[i] = whatReplace
          found = True
          print("Found")
          print(lists)
    else:
      print("cannot be found")

  accountNames[i] = lists
  file = open(fileRaw, "w")
  for i in range(0, len(lists)):
    if i == len(lists) - 1:
      file.write(lists[i])
    else:
      file.write(lists[i] + "\n")
  file.close()



def deleteSearchTerm(fileRaw, i):
  file = open(fileRaw, "r")
  
  reading = file.read()
  lists = reading.split("\n")
  file.close()
  print(lists)
  numberInLists = len(lists)
  global Tyrell, Illingworth, Gale_Close39, Silbury, Flanders54, Mount27, Sheldrick, North_Place, Garendon
  found = False
  while found == False:
    print("\n\nWhat would you like to delete?")
    searchDelete = input("Tenant Name/Company ID: ")
    if searchDelete in lists:
      for i in range(0, len(lists)):
        if lists[i] == searchDelete:
          lists.pop(i)
          found = True
          print("Found")
          print(lists)
    else:
      print("\n\nCannot be found")

  accountNames[i] = lists
  file = open(fileRaw, "w")
  for i in range(0, len(lists)):
    if i == len(lists) - 1:
      file.write(lists[i])
    else:
      file.write(lists[i] + "\n")
  file.close()


  
def addSearchTerm(fileRaw, i):
  file = open(fileRaw, "r")
  
  reading = file.read()
  lists = reading.split("\n")
  file.close()
  print(lists)
  numberInLists = len(lists)
  global Tyrell, Illingworth, Gale_Close39, Silbury, Flanders54, Mount27, Sheldrick, North_Place, Garendon
  found = False
  while found == False:
    sure = False
    while sure == False:
      whatAdd = input("\nWhat would you like to add?\nMake sure that it is correct and will be able to work for the next month\nTenant Name/Company ID: ")
      checky = False
      while checky == False:
        RUSure = input("\nAre you sure you want " + str(whatAdd) + " to be added?\nYes/y or No/n: ")
        if RUSure == "Yes" or RUSure == "yes" or RUSure == "y":
          lists.append(whatAdd)
          sure = True
          checky = True
        elif RUSure == "No" or RUSure == "no" or RUSure == "n":
          sure = True
          checky = True
          print("\nOK")
        else:
          sure = False
          checky = False
          print("\nSorry, I did not understand.")

    found = True
    print(lists)
    
    
  accountNames[i] = lists
  file = open(fileRaw, "w")
  for i in range(0, len(lists)):
    if i == len(lists) - 1:
      file.write(lists[i])
    else:
      file.write(lists[i] + "\n")
  file.close()



# -------------------------------------------------------------=======[[[[[[end of functions]]]]]]=======-----------------------------------------------------------------
# -------------------------------------------------------------=======[[[[[[end of functions]]]]]]=======-----------------------------------------------------------------


#show: date, account number (bank account), amount +/- and
#who it belongs to ((+)who is paying x or (-) who x is paying) and
#create new for property and colour.
itr()



mainloop = True

while mainloop == True:

  #Welcome

  print("\n\n\n--------------------------------------------------------------------------------\n\nWelcome, what would you like to do right now? \n\n\t  ______________________________________  \n\t |Sort (keyletter S/s)          | \n\t |Change Tenant list/Tenant Number (C/c)| \n\t |Add Tenant Name/ Company ID (A/a)   | \n\t |Delete Tenant Name/ Company ID (D/d)  | \n\t |End process (E/e)           |\n\t |______________________________________|\n\n")

  User_input = input("s, c, a, d or e: ")

          
  if (User_input == "S") or (User_input == "s"):

    
    # name of Excel file
    excelFileName = input("\n\nWhat do you want the output file to be called?\n(Make sure that you do not use special characters. Definately not fullstops)\nFile name: ")
    fileNameXlsx = str(excelFileName) + ".xlsx"
    
    # reminders
    print("\n\nCopy the lines needed into a fresh excel sheet and move it into the same folder as this program.")
    
    sheetName = (input("\nWhat is the name of the file?\nFile name: ")) + ".xlsx"

    linesInSheet = int(input("\n\nEnter amount of lines in excel sheet\nLines: "))
    linesActual = linesInSheet - 1 #account for top line


    #initialise workbooks + sheets
    book = openpyxl.load_workbook(sheetName)
    ws = book.active
##    wsprop = ws.sheet_properties

    readColumnF(ws)

    #Save the sheet
    book.save(filename = fileNameXlsx)



  elif User_input == "C" or User_input == "c":
    check = False
    while check == False:
      print("Which property would you like to change?\ne.g.: " + str(accountNamesString))
      propChange = input("Property: ") + "F"
      if propChange in accountSaveStringsLong:
        print("Found")
        numberOfSaves = len(accountSaveFiles)
        for i in range(0, numberOfSaves):
          if propChange == accountSaveStrings[i]:
            replaceSearchTerm(accountSaveFiles[i], i)
            check = True
      else:
        print("Sorry, I cannot find that property. Please write it like this:\n" + str(accountNamesString))
##    mainloop = False


  elif User_input == "A" or User_input == "a":
    check = False
    while check == False:
      print("Which property would you like to change?\ne.g.: " + str(accountNamesString))
      propChange = input("Property: ") + "F"
      if propChange in accountSaveStringsLong:
        print("Found")
        numberOfSaves = len(accountSaveFiles)
        for i in range(0, numberOfSaves):
          if propChange == accountSaveStrings[i]:
            addSearchTerm(accountSaveFiles[i], i)
            check = True
      else:
        print("Sorry, I cannot find that property. Please write it like this:\n" + str(accountNamesString))
##    mainloop = False


  elif User_input == "D" or User_input == "d":
    check = False
    while check == False:
      print("Which property would you like to change?\ne.g. " + str(accountNamesString))
      propChange = input("Property: ") + "F"
      if propChange in accountSaveStringsLong:
        print("Found")
        numberOfSaves = len(accountSaveFiles)
        for i in range(0, numberOfSaves):
          if propChange == accountSaveStrings[i]:
            deleteSearchTerm(accountSaveFiles[i], i)
            check = True
      else:
        print("Sorry, I cannot find that property. Please write it like this:\n" + str(accountNamesString))
##    mainloop = False


  elif User_input == "E" or User_input == "e":
    print("\n\n\nThank you for using Rohan & Co( - 'Co' + 'Keanu') Services.\n:D")
    mainloop = False


  else:
    print("Please try again, your input was '" + str(User_input) + "'\n\n\n")






##Patch notes
##  - col I for extra info
##  - finish col J -Room info (number, name, ID)
##  - changed order of Cols H, I, J
##  - manual notes now are in yellow
##  - increased length of colour line(A to H now)
##  - column headings

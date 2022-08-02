import os
from openpyxl.styles import PatternFill


def getSaveFiles(saveFileFolder):
  TyrellF = os.path.join(saveFileFolder, "SaveTyrellFile.txt")
  IllingworthF = os.path.join(saveFileFolder, "SaveIllingworthFile.txt")
  Gale_Close39F = os.path.join(saveFileFolder, "SaveGale_CloseFile.txt")
  SilburyF = os.path.join(saveFileFolder, "SaveSilburyFile.txt")
  Flanders54F = os.path.join(saveFileFolder, "SaveFlanders54File.txt")
  Mount27F = os.path.join(saveFileFolder, "SaveMount27File.txt")
  SheldrickF = os.path.join(saveFileFolder, "SaveSheldrickFile.txt")
  North_PlaceF = os.path.join(saveFileFolder, "SaveNorth_PlaceFile.txt")
  GarendonF = os.path.join(saveFileFolder, "SaveGarendonFile.txt")

  openTyrellF = open(TyrellF, "r")
  openIllingworthF = open(IllingworthF, "r")
  openGaleCloseF = open(Gale_Close39F, "r")
  openSilburyF = open(SilburyF, "r")
  openFlanders54F = open(Flanders54F, "r")
  openMount27F = open(Mount27F, "r")
  openSheldrickF = open(SheldrickF, "r")
  openNorthPlaceF = open(North_PlaceF, "r")
  openGarendonF = open(GarendonF, "r")

  readerTyrellF = openTyrellF.read()
  readerIllingworthF = openIllingworthF.read()
  readerGaleCloseF = openGaleCloseF.read()
  readerSilburyF = openSilburyF.read()
  readerFlanders54F = openFlanders54F.read()
  readerMount27F = openMount27F.read()
  readerSheldrickF = openSheldrickF.read()
  readerNorthPlaceF = openNorthPlaceF.read()
  readerGarendonF = openGarendonF.read()

  Tyrell = readerTyrellF.split("\n")
  Illingworth = readerIllingworthF.split("\n")
  Gale_Close39 = readerGaleCloseF.split("\n")
  Silbury = readerSilburyF.split("\n")
  Flanders54 = readerFlanders54F.split("\n")
  Mount27 = readerMount27F.split("\n")
  Sheldrick = readerSheldrickF.split("\n")
  North_Place = readerNorthPlaceF.split("\n")
  Garendon = readerGarendonF.split("\n")

  openTyrellF.close()
  openIllingworthF.close()
  openGaleCloseF.close()
  openSilburyF.close()
  openFlanders54F.close()
  openMount27F.close()
  openSheldrickF.close()
  openNorthPlaceF.close()
  openGarendonF.close()

  accountSaveFiles = [TyrellF, IllingworthF, Gale_Close39F, SilburyF, Flanders54F, Mount27F, SheldrickF, North_PlaceF, GarendonF]

  accountNames = [Tyrell, Illingworth, Gale_Close39, Silbury, Flanders54, Mount27, Sheldrick, North_Place, Garendon]

  roomSaveFiles2 = [TyrellF, IllingworthF, Gale_Close39F, SilburyF, Flanders54F, Mount27F, SheldrickF, North_PlaceF, GarendonF]

  return accountSaveFiles, accountNames, roomSaveFiles2

def getRoomFiles(roomFileFolder):
  join = [lambda y, x : os.path.join(x, y)]
  
  TyrellS = join[0]("RoomTyrell.txt", roomFileFolder)
  IllingworthS = join[0]("RoomIllingworth.txt", roomFileFolder)
  Gale_Close39S = join[0]("RoomGale_Close.txt", roomFileFolder)
  SilburyS = join[0]("RoomSilbury.txt", roomFileFolder)
  Flanders54S = join[0]("RoomFlanders54.txt", roomFileFolder)
  Mount27S = join[0]("RoomMount27.txt", roomFileFolder)
  SheldrickS = join[0]("RoomSheldrick.txt", roomFileFolder)
  North_PlaceS = join[0]("RoomNorth_Place.txt", roomFileFolder)
  GarendonS = join[0]("RoomGarendon.txt", roomFileFolder)

  openTyrellR = open(TyrellS, "r")
  openIllingworthR = open(IllingworthS, "r")
  openGaleCloseR = open(Gale_Close39S, "r")
  openSilburyR = open(SilburyS, "r")
  openFlanders54R = open(Flanders54S, "r")
  openMount27R = open(Mount27S, "r")
  openSheldrickR = open(SheldrickS, "r")
  openNorthPlaceR = open(North_PlaceS, "r")
  openGarendonR = open(GarendonS, "r")

  readerTyrellR = openTyrellR.read()
  readerIllingworthR = openIllingworthR.read()
  readerGaleCloseR = openGaleCloseR.read()
  readerSilburyR = openSilburyR.read()
  readerFlanders54R = openFlanders54R.read()
  readerMount27R = openMount27R.read()
  readerSheldrickR = openSheldrickR.read()
  readerNorthPlaceR = openNorthPlaceR.read()
  readerGarendonR = openGarendonR.read()

  TyrellR = readerTyrellR.split("\n")
  IllingworthR = readerIllingworthR.split("\n")
  Gale_Close39R = readerGaleCloseR.split("\n")
  SilburyR = readerSilburyR.split("\n")
  Flanders54R = readerFlanders54R.split("\n")
  Mount27R = readerMount27R.split("\n")
  SheldrickR = readerSheldrickR.split("\n")
  North_PlaceR = readerNorthPlaceR.split("\n")
  GarendonR = readerGarendonR.split("\n")

  openTyrellR.close()
  openIllingworthR.close()
  openGaleCloseR.close()
  openSilburyR.close()
  openFlanders54R.close()
  openMount27R.close()
  openSheldrickR.close()
  openNorthPlaceR.close()
  openGarendonR.close()

  roomSaveFiles = [TyrellR, IllingworthR, Gale_Close39R, SilburyR, Flanders54R, Mount27R, SheldrickR, North_PlaceR, GarendonR]

  roomSaveFilesLocs = [TyrellS, IllingworthS, Gale_Close39S, SilburyS, Flanders54S, Mount27S, SheldrickS, North_PlaceS, GarendonS]

  return roomSaveFiles, roomSaveFilesLocs

def getAll(saveFileFolder, roomFileFolder):
  NotesMemo = os.path.join("Notes", "NotesMemo.txt")
  NotesNote = os.path.join("Notes",  "NotesNote.txt")

  investorsF = os.path.join("investors.txt")

  openNotesMemo = open(NotesMemo, "r")
  openNotesNote = open(NotesNote, "r")

  openInvestors = open(investorsF, "r")

  readerMemo = openNotesMemo.read()
  readerNote = openNotesNote.read()

  readerInvestors = openInvestors.read()

  Memo = readerMemo.split("\n")
  Note = readerNote.split("\n")

  investors = readerInvestors.split("\n")

  openNotesMemo.close()
  openNotesNote.close()

  # full line is income, just the property is outgo

  Dgreen = PatternFill(start_color='139a20', end_color='139a20', fill_type='solid')
  Dblue = PatternFill(start_color='5077f7', end_color='5077f7', fill_type='solid')
  marron = PatternFill(start_color='850a1c', end_color='850a1c', fill_type='solid')
  purpink = PatternFill(start_color='cd8070', end_color='cd8070', fill_type='solid')
  orange = PatternFill(start_color='f6b723', end_color='f6b723', fill_type='solid')
  Lblue = PatternFill(start_color='269bdf', end_color='269bdf', fill_type='solid')
  lpurp = PatternFill(start_color='a692c3', end_color='a692c3', fill_type='solid')
  vibpink = PatternFill(start_color='d725e9', end_color='d725e9', fill_type='solid')
  peach = PatternFill(start_color='ecc18e', end_color='ecc18e', fill_type='solid')
  redish = PatternFill(start_color='B25651', end_color='B25651', fill_type='solid')
  none = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid') #actually yellow colour

  roomSaveFiles, roomSaveFilesLoc = getRoomFiles(roomFileFolder)
  accountSaveFiles, accountNames, roomSaveFiles2 = getSaveFiles(saveFileFolder)

  accountSaveStrings = ["TyrrellF", "IllingworthF", "39 Gale CloseF", "SilburyF", "54 FlandersF", "27 MountF", "SheldrickF", "North_PlaceF", "GarendonF"]
  accountSaveStringsLong = "TyrrellF IllingworthF 39 Gale CloseF SilburyF 54 FlandersF 27 MountF SheldrickF North_PlaceF GarendonF"
  accountNamesString = ["Tyrrell", "Illingworth", "39 Gale Close", "Silbury", "54 Flanders", "27 Mount", "Sheldrick", "North_Place", "Garendon"]

  accountColoursFill = [Dgreen, Dblue, redish, purpink, orange, Lblue, lpurp, vibpink, peach]

  return accountSaveFiles, roomSaveFiles, accountSaveStrings, accountSaveStringsLong, accountNamesString, accountNames, accountColoursFill, Memo, Note, none, NotesNote, NotesMemo, investors
